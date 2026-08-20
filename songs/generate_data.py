#!/usr/bin/env python3
"""
读取 data.xlsx，生成 data.js

用法：
  python generate_data.py

XLSX 包含三个 Sheet：

Sheet 1 - "songs"（歌曲信息）：
  song_name    - 歌曲名称（中文）
  song_name_jp - 歌曲名称（日文）
  album        - 所属专辑
  album_year   - 专辑年份
  release_date - 发行日期（YYYY/M/D）
  cover        - 封面图片路径
  type         - 原创 / 翻唱
  lyricist     - 作词
  composer     - 作曲
  arranger     - 编曲
  first_stage  - 首次登台
  mv_url       - MV 链接（可选）
  lyrics_jp    - 日文歌词（\\n 换行）
  lyrics_cn    - 中文歌词（\\n 换行）
  appearances  - 收录履历（逗号分隔）

Sheet 2 - "comments"（制作人讲解，按歌曲名关联，可多条）：
  song_name      - 歌曲名称（与 Sheet1 对应）
  comment_text   - 讲解内容
  comment_source - 出处链接（可选）
  comment_from   - 讲解来源标识（如"Diggy-MO'"、"佐佐木李子"等，用于区分多条讲解）

Sheet 3 - "live_history"（Live 演唱履历，按歌曲名关联）：
  song_name    - 歌曲名称（与 Sheet1 对应）
  live_date    - Live 日期
  live_venue   - Live 场地
  live_name    - Live 名称
  has_video    - 有无影像（yes / no）
  video_url    - 影像链接（has_video=yes 时填写）
"""

import os
import re
from collections import OrderedDict

import openpyxl

OUTPUT_FILE = "data.js"
XLSX_FILE = "data.xlsx"

# ---------- 如果 Excel 不存在，自动创建模板 ----------
if not os.path.exists(XLSX_FILE):
    wb = openpyxl.Workbook()
    ws_songs = wb.active
    ws_songs.title = "songs"
    ws_songs.append([
        "song_name", "song_name_jp", "album", "album_year", "release_date", "cover",
        "type", "lyricist", "composer", "arranger", "first_stage", "mv_url",
        "lyrics_jp", "lyrics_cn", "appearances"
    ])
    ws_songs.append([
        "黑色生日", "黒の誕生日", "Alea jacta est", "2024", "2024/5/15", "images/黑色生日.png",
        "原创", "Diggy-MO'", "Diggy-MO'", "Diggy-MO'", "2024/6/8 Ave Mujica 1st Live", "https://example.com/mv",
        "歌词第一行\\n歌词第二行", "中文歌词第一行\\n中文歌词第二行",
        "Alea jacta est, 精选集"
    ])

    ws_comments = wb.create_sheet("comments")
    ws_comments.append(["song_name", "comment_text", "comment_source", "comment_from"])
    ws_comments.append(["黑色生日", "这是 Ave Mujica 的首张单曲主打歌，由……", "https://example.com/interview", "Diggy-MO'"])

    ws_live = wb.create_sheet("live_history")
    ws_live.append(["song_name", "live_date", "live_venue", "live_name", "has_video", "video_url"])
    ws_live.append(["黑色生日", "2024/6/8", "Tokyo Dome", "Ave Mujica 1st Live", "yes", "https://example.com"])
    ws_live.append(["黑色生日", "2024/6/30", "Nagoya", "Live Tour Nagoya", "no", ""])

    wb.save(XLSX_FILE)
    print(f"已创建模板文件: {XLSX_FILE}")
    print("请填写数据后重新运行本脚本。")
    exit(0)

# ---------- 读取 Excel ----------
wb = openpyxl.load_workbook(XLSX_FILE)

def read_sheet(sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(header):
            if h:
                d[h.lower()] = str(row[i]).strip() if row[i] is not None and str(row[i]).strip() != "None" else ""
        if any(v for v in d.values()):
            result.append(d)
    return result

songs_raw = read_sheet("songs")
comments_raw = read_sheet("comments")
live_raw = read_sheet("live_history")

print(f"从 data.xlsx 读取了 {len(songs_raw)} 首歌曲, {len(comments_raw)} 条 Comment, {len(live_raw)} 条 Live 记录")

# ---------- 按 song_name 合并 Comment 记录 ----------
comments_map = {}
for cr in comments_raw:
    sn = cr.get("song_name", "")
    if not sn:
        continue
    if sn not in comments_map:
        comments_map[sn] = []
    entry = {
        "text": cr.get("comment_text", ""),
        "source": cr.get("comment_source", ""),
        "from": cr.get("comment_from", "")
    }
    comments_map[sn].append(entry)

# ---------- 按 song_name 合并 Live 记录 ----------
def normalize_date(raw):
    raw = raw.strip()
    if not raw:
        return raw
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})\s", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})$", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    return raw

def fix_path(p):
    """将 images/ 路径转为 ../images/（HTML 在子文件夹中）"""
    p = p.strip()
    if p.startswith("images/"):
        return "../" + p
    return p

# ---------- 按 song_name 合并 Live 记录 ----------
live_map = {}
for lr in live_raw:
    sn = lr.get("song_name", "")
    if not sn:
        continue
    if sn not in live_map:
        live_map[sn] = []
    entry = {
        "date": normalize_date(lr.get("live_date", "")),
        "venue": lr.get("live_venue", ""),
        "name": lr.get("live_name", ""),
        "has_video": lr.get("has_video", "").lower() == "yes",
        "video_url": lr.get("video_url", "")
    }
    live_map[sn].append(entry)

# ---------- 处理歌曲数据 ----------
songs = []
for s in songs_raw:
    sn = s.get("song_name", "")
    if not sn:
        continue
    song = {
        "name": sn,
        "name_jp": s.get("song_name_jp", ""),
        "album": s.get("album", ""),
        "album_year": s.get("album_year", ""),
        "release_date": normalize_date(s.get("release_date", "")),
        "cover": fix_path(s.get("cover", "")),
        "type": s.get("type", ""),
        "lyricist": s.get("lyricist", ""),
        "composer": s.get("composer", ""),
        "arranger": s.get("arranger", ""),
        "first_stage": s.get("first_stage", ""),
        "mv_url": s.get("mv_url", ""),
        "lyrics_jp": s.get("lyrics_jp", ""),
        "lyrics_cn": s.get("lyrics_cn", ""),
        "appearances": [a.strip() for a in s.get("appearances", "").split(",") if a.strip()],
        "comments": comments_map.get(sn, []),
        "live_history": live_map.get(sn, [])
    }
    songs.append(song)

import json

# ---------- 生成 data.js ----------
def js_str(s):
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "")

lines = []
lines.append("// 曲目数据")
lines.append("// 由 generate_data.py 自动生成，请勿手动修改")
lines.append("// 编辑 data.xlsx 后运行「一键更新数据.bat」即可更新")
lines.append("")
lines.append("const songsData = [")

for i, song in enumerate(songs):
    lines.append("  {")
    lines.append(f'    name: "{js_str(song["name"])}",')
    lines.append(f'    name_jp: "{js_str(song["name_jp"])}",')
    lines.append(f'    album: "{js_str(song["album"])}",')
    lines.append(f'    album_year: "{js_str(song["album_year"])}",')
    lines.append(f'    release_date: "{js_str(song["release_date"])}",')
    lines.append(f'    cover: "{js_str(song["cover"])}",')
    lines.append(f'    type: "{js_str(song["type"])}",')
    lines.append(f'    lyricist: "{js_str(song["lyricist"])}",')
    lines.append(f'    composer: "{js_str(song["composer"])}",')
    lines.append(f'    arranger: "{js_str(song["arranger"])}",')
    lines.append(f'    first_stage: "{js_str(song["first_stage"])}",')
    lines.append(f'    mv_url: "{js_str(song["mv_url"])}",')
    lines.append(f'    lyrics_jp: "{js_str(song["lyrics_jp"])}",')
    lines.append(f'    lyrics_cn: "{js_str(song["lyrics_cn"])}",')
    # 使用 json.dumps 安全输出 appearances 列表
    lines.append(f'    appearances: {json.dumps(song["appearances"], ensure_ascii=False)},')
    # comments
    if song["comments"]:
        lines.append("    comments: [")
        for j, c in enumerate(song["comments"]):
            lines.append("      {")
            lines.append(f'        text: "{js_str(c["text"])}",')
            lines.append(f'        source: "{js_str(c["source"])}",')
            lines.append(f'        from: "{js_str(c["from"])}"')
            lines.append("      }" + ("," if j < len(song["comments"]) - 1 else ""))
        lines.append("    ],")
    else:
        lines.append("    comments: [],")
    # live_history
    if song["live_history"]:
        lines.append("    live_history: [")
        for j, lh in enumerate(song["live_history"]):
            lines.append("      {")
            lines.append(f'        date: "{js_str(lh["date"])}",')
            lines.append(f'        venue: "{js_str(lh["venue"])}",')
            lines.append(f'        name: "{js_str(lh["name"])}",')
            lines.append(f'        has_video: {"true" if lh["has_video"] else "false"},')
            lines.append(f'        video_url: "{js_str(lh["video_url"])}"')
            lines.append("      }" + ("," if j < len(song["live_history"]) - 1 else ""))
        lines.append("    ]")
    else:
        lines.append("    live_history: []")
    lines.append("  }" + ("," if i < len(songs) - 1 else ""))

lines.append("];")

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")

print(f"Done! {len(songs)} songs written to {OUTPUT_FILE}")