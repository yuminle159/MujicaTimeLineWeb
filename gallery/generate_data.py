#!/usr/bin/env python3
"""
读取 data.xlsx，生成 data.js 用于画廊展示

用法：
  py generate_data.py

XLSX 包含一个 Sheet - "images"：
  filename    - 图片路径（相对于 gallery/ 目录，如 ../images/xxx.webp）
  title       - 图片标题
  date        - 日期（YYYY/M/D 或 文字描述）
  tags        - 标签（两层格式：CATEGORY:subtag，逗号分隔）
                例如 "LIVE:0th, LIVE:1st, SINGLE:cover"
                分类名在前，子标签用冒号分隔，不同分类用逗号分隔
  description - 描述文字（可选）
"""

import os
import json

import openpyxl

OUTPUT_FILE = "data.js"
XLSX_FILE = "data.xlsx"

# ---------- 如果 Excel 不存在，自动创建模板 ----------
if not os.path.exists(XLSX_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "images"
    ws.append(["filename", "title", "date", "tags", "description"])
    ws.append([
        "../images/0thlive返图1.webp",
        "0th Live 返图",
        "2023/6/4",
        "LIVE:0th",
        "Ave Mujica 0th Live 舞台返图"
    ])
    ws.append([
        "../images/1st live kv.webp",
        "1st Live 主视觉",
        "2023/6/8",
        "LIVE:1st, LIVE:kv",
        "Perdere Omnia 主视觉海报"
    ])
    ws.append([
        "../images/黑色生日.webp",
        "黑色生日",
        "2024/5/15",
        "SINGLE:cover",
        "1st Single 封面"
    ])
    ws.append([
        "../images/双月.webp",
        "双月",
        "2024/10/12",
        "SINGLE:cover",
        "2nd Single 封面"
    ])

    wb.save(XLSX_FILE)
    print(f"已创建模板文件: {XLSX_FILE}")
    print("请填写数据后重新运行本脚本。")
    exit(0)

# ---------- 读取 Excel ----------
wb = openpyxl.load_workbook(XLSX_FILE)

def read_sheet(sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(header):
            if h:
                d[h.lower()] = str(row[i]).strip() if row[i] is not None and str(row[i]).strip() != "None" else ""
        if any(v for v in d.values()):
            result.append(d)
    return result

images_raw = read_sheet("images")
print(f"从 data.xlsx 读取了 {len(images_raw)} 张图片")

# ---------- 解析两层 Tag 格式 ----------
def parse_tags(raw_tags):
    """
    解析 "LIVE:0th, LIVE:1st, SINGLE:cover" 为嵌套 dict
    { "LIVE": ["0th", "1st"], "SINGLE": ["cover"] }
    """
    result = {}
    if not raw_tags:
        return result
    for tag in raw_tags.split(","):
        tag = tag.strip()
        if not tag:
            continue
        if ":" in tag:
            cat, sub = tag.split(":", 1)
            cat = cat.strip()
            sub = sub.strip()
            if sub:
                if cat not in result:
                    result[cat] = []
                if sub not in result[cat]:
                    result[cat].append(sub)
        else:
            # 无子标签的分类（如纯 "LIVE"）
            cat = tag.strip()
            if cat and cat not in result:
                result[cat] = []
    return result

# ---------- 处理数据 ----------
images = []
for img in images_raw:
    if not img.get("filename"):
        continue
    item = {
        "filename": img.get("filename", ""),
        "title": img.get("title", ""),
        "date": img.get("date", ""),
        "tags": parse_tags(img.get("tags", "")),
        "description": img.get("description", "")
    }
    images.append(item)

# ---------- 生成 data.js ----------
def js_str(s):
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "")

lines = []
lines.append("// 画廊图片数据")
lines.append("// 由 generate_data.py 自动生成，请勿手动修改")
lines.append("// 编辑 data.xlsx 后运行「一键更新数据.bat」即可更新")
lines.append("// tags 格式: { CATEGORY: [subtag, ...], ... }")
lines.append("")
lines.append("const galleryData = [")

for i, item in enumerate(images):
    lines.append("  {")
    lines.append(f'    filename: "{js_str(item["filename"])}",')
    lines.append(f'    title: "{js_str(item["title"])}",')
    lines.append(f'    date: "{js_str(item["date"])}",')
    # 输出嵌套 tags
    tags_json = json.dumps(item["tags"], ensure_ascii=False)
    lines.append(f'    tags: {tags_json},')
    lines.append(f'    description: "{js_str(item["description"])}"')
    lines.append("  }" + ("," if i < len(images) - 1 else ""))

lines.append("];")

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")

print(f"Done! {len(images)} images written to {OUTPUT_FILE}")