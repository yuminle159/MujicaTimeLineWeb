#!/usr/bin/env python3
"""
读取 data.xlsx，生成 data.js

用法：
  python generate_data.py

XLSX 包含三个 Sheet：

Sheet 1 - "lives"（演唱会信息）：
  live_name    - 演唱会名称
  live_date    - 日期（YYYY-MM-DD 或 YYYY/M/D）
  live_venue   - 场地
  poster       - 海报图片路径
  kv           - KV 横图路径（可选，点击可放大全图）
  video_url    - 影像链接（可选）
  description  - 简介（可选）

Sheet 2 - "setlist"（曲目列表，按演唱会名关联）：
  live_name       - 演唱会名称（与 Sheet1 对应）
  track_num       - 曲目序号（如 "01", "M0", "SE" 等，留空显示占位符）
  track_title     - 歌曲名
  highlight_label - 高亮标签（如 "Acoustic Arrange"，可选）
  highlight_text  - 高亮说明（可选）

Sheet 3 - "backstage"（幕后照片，按演唱会名关联）：
  live_name    - 演唱会名称（与 Sheet1 对应）
  photo        - 照片路径
  credit       - 来源署名（如 "@Cast_Vocal"）
  credit_text  - 来源说明文字
  source_url   - 来源链接（可选）
  source_label - 来源链接标签（如 "View on X"，可选）
  group_id     - 分组 ID（同一值的照片归入同一轮播组，留空 = 独立显示）
"""

import os
import re
import json
from collections import OrderedDict

import openpyxl

OUTPUT_FILE = "data.js"
XLSX_FILE = "data.xlsx"

# ---------- 如果 Excel 不存在，自动创建模板 ----------
if not os.path.exists(XLSX_FILE):
    wb = openpyxl.Workbook()
    ws_lives = wb.active
    ws_lives.title = "lives"
    ws_lives.append(["live_name", "live_date", "live_venue", "poster", "kv", "video_url", "description"])
    ws_lives.append([
        "Ave Mujica 1st LIVE「Perdere Omnia」",
        "2024/1/27",
        "横浜1000CLUB",
        "../images/live_1st.jpg",
        "../images/live_1st.jpg",
        "https://example.com/video",
        ""
    ])
    ws_lives.append([
        "Ave Mujica 2nd LIVE「Quaerere Lumina」",
        "2024/7/7",
        "愛知県芸術劇場 大ホール",
        "../images/live_2nd.jpg",
        "../images/live_2nd.jpg",
        "",
        "爱知公演初日"
    ])

    ws_setlist = wb.create_sheet("setlist")
    ws_setlist.append(["live_name", "track_num", "track_title", "highlight_label", "highlight_text", "mc_file"])
    ws_setlist.append(["Ave Mujica 1st LIVE「Perdere Omnia」", "01", "Ave Mujica", "", ""])
    ws_setlist.append(["Ave Mujica 1st LIVE「Perdere Omnia」", "02", "黒の誕生日", "Live Highlight", "首次公开演出，全场暗转后以钢琴独奏开场"])

    ws_backstage = wb.create_sheet("backstage")
    ws_backstage.append(["live_name", "photo", "credit", "credit_text", "source_url", "source_label", "group_id"])
    ws_backstage.append([
        "Ave Mujica 1st LIVE「Perdere Omnia」",
        "../images/backstage_1.jpg",
        "@Official_Info",
        "初日公演圆满结束！",
        "https://x.com/example",
        "View on X",
        ""
    ])

    wb.save(XLSX_FILE)
    print(f"已创建模板文件: {XLSX_FILE}")
    print("请填写数据后重新运行本脚本。")
    exit(0)

# ---------- 读取 Excel ----------
wb = openpyxl.load_workbook(XLSX_FILE)

def read_sheet(sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(header):
            if h:
                d[h.lower()] = str(row[i]).strip() if row[i] is not None and str(row[i]).strip() != "None" else ""
        if any(v for v in d.values()):
            result.append(d)
    return result

lives_raw = read_sheet("lives")
setlist_raw = read_sheet("setlist")
backstage_raw = read_sheet("backstage")

print(f"从 data.xlsx 读取了 {len(lives_raw)} 场演唱会, {len(setlist_raw)} 条曲目, {len(backstage_raw)} 条幕后照片")

# ---------- 日期标准化 ----------
def normalize_date(raw):
    raw = raw.strip()
    if not raw:
        return raw
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})\s", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})$", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})\s", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    return raw

def fix_path(p):
    """将 images/ 路径转为 ../images/（HTML 在子文件夹中）"""
    p = p.strip()
    if p.startswith("images/"):
        return "../" + p
    return p

# ---------- 按 live_name 合并 Setlist ----------
setlist_map = {}
for sr in setlist_raw:
    sn = sr.get("live_name", "")
    if not sn:
        continue
    if sn not in setlist_map:
        setlist_map[sn] = []
    setlist_map[sn].append({
        "num": sr.get("track_num", ""),
        "title": sr.get("track_title", ""),
        "highlight_label": sr.get("highlight_label", ""),
        "highlight_text": sr.get("highlight_text", ""),
        "mc_file": sr.get("mc_file", ""),       # ← 新增：md 文件路径
    })

# ---------- 按 live_name 合并 Backstage ----------
backstage_map = {}
for br in backstage_raw:
    sn = br.get("live_name", "")
    if not sn:
        continue
    if sn not in backstage_map:
        backstage_map[sn] = []
    backstage_map[sn].append({
        "photo": fix_path(br.get("photo", "")),
        "credit": br.get("credit", ""),
        "credit_text": br.get("credit_text", ""),
        "source_url": br.get("source_url", ""),
        "source_label": br.get("source_label", ""),
        "group_id": br.get("group_id", "")
    })

# ---------- 处理演唱会数据 ----------
lives = []
for l in lives_raw:
    name = l.get("live_name", "")
    if not name:
        continue
    live = {
        "name": name,
        "date": normalize_date(l.get("live_date", "")),
        "venue": l.get("live_venue", ""),
        "poster": fix_path(l.get("poster", "")),
        "kv": fix_path(l.get("kv", "")),
        "video_url": l.get("video_url", ""),
        "description": l.get("description", ""),
        "setlist": setlist_map.get(name, []),
        "backstage": backstage_map.get(name, [])
    }
    lives.append(live)

# ---------- 读取 MC 内容（通过 setlist 中的 mc_file 字段指定 md 文件） ----------
# mc_file 为相对于 live/ 目录的路径，例如 "mc/0_0.md"
LIVE_DIR = os.path.dirname(__file__)
mc_count = 0
for live in lives:
    for track in live.get("setlist", []):
        mc_file = track.get("mc_file", "").strip()
        if mc_file:
            md_path = os.path.join(LIVE_DIR, mc_file)
            if os.path.isfile(md_path):
                with open(md_path, "r", encoding="utf-8") as f:
                    track["mc_content"] = f.read().strip()
                mc_count += 1
            else:
                print(f"  [WARN] MC 文件不存在: {mc_file}")
if mc_count:
    print(f"已加载 {mc_count} 个 MC 内容文件")

# ---------- 生成 data.js ----------
def js_str(s):
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "")

lines = []
lines.append("// 演唱会数据")
lines.append("// 由 generate_data.py 自动生成，请勿手动修改")
lines.append("// 编辑 data.xlsx 后运行「一键更新数据.bat」即可更新")
lines.append("")
lines.append("const livesData = [")

for i, live in enumerate(lives):
    lines.append("  {")
    lines.append(f'    name: "{js_str(live["name"])}",')
    lines.append(f'    date: "{js_str(live["date"])}",')
    lines.append(f'    venue: "{js_str(live["venue"])}",')
    lines.append(f'    poster: "{js_str(live["poster"])}",')
    lines.append(f'    kv: "{js_str(live["kv"])}",')
    lines.append(f'    video_url: "{js_str(live["video_url"])}",')
    lines.append(f'    description: "{js_str(live["description"])}",')
    # setlist
    lines.append(f'    setlist: {json.dumps(live["setlist"], ensure_ascii=False)},')
    # backstage
    lines.append(f'    backstage: {json.dumps(live["backstage"], ensure_ascii=False)}')
    lines.append("  }" + ("," if i < len(lives) - 1 else ""))

lines.append("];")

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")

print(f"Done! {len(lives)} lives written to {OUTPUT_FILE}")