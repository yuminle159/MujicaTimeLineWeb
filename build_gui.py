"""
 wijipedia 数据更新工具
 统一数据生成 + 图片转 WebP
"""
import os
import sys
import io
import tkinter as tk
from tkinter import ttk, messagebox

try:
    from PIL import Image
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

# PyInstaller --onefile 会将数据文件解压到临时目录
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
    PROJECT_DIR = os.path.dirname(sys.executable)  # exe 所在目录（用于输出 JS 文件）
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    PROJECT_DIR = BASE_DIR

# 模块名称映射
MODULE_NAMES = {
    "announcements":     "公告",
    "songs":             "歌曲（Songs）",
    "lives":             "演唱会（Live）",
    "timeline":          "时间线（Timeline）",
    "gallery_images":    "画廊（Gallery）",
}


def convert_images_to_webp(folder_path, log_func=None, skip_existing_webp=True):
    """将文件夹中的图片转为 WebP 格式并压缩（max_width=1920, quality=75, method=6）
    :param skip_existing_webp: 如果为 True，跳过已存在同名 .webp 的文件
    """
    def log(msg):
        (log_func or print)(msg)

    if not os.path.isdir(folder_path):
        log(f"  [跳过] 目录不存在: {folder_path}")
        return 0

    supported_formats = ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tiff', '.tif')
    converted = 0
    for filename in os.listdir(folder_path):
        if not filename.lower().endswith(supported_formats):
            continue

        file_path = os.path.join(folder_path, filename)
        name_without_ext = os.path.splitext(filename)[0]
        webp_path = os.path.join(folder_path, f"{name_without_ext}.webp")

        # 跳过已存在 webp 的文件
        if skip_existing_webp and os.path.exists(webp_path):
            continue

        temp_path = webp_path + ".tmp"
        orig_size_kb = os.path.getsize(file_path) / 1024 if os.path.exists(file_path) else 0

        try:
            with Image.open(file_path) as img:
                if img.mode not in ('RGB', 'RGBA'):
                    img = img.convert('RGBA')

                # 缩放
                max_width = 1920
                if img.width > max_width:
                    new_h = int(img.height * (max_width / img.width))
                    img = img.resize((max_width, new_h), Image.Resampling.LANCZOS)
                    log(f"    [缩放] {filename}: {img.width}x{img.height} -> {max_width}x{new_h}")

                img.save(temp_path, 'webp', quality=75, method=6)

            new_size_kb = os.path.getsize(temp_path) / 1024
            os.replace(temp_path, webp_path)
            log(f"    {filename} ({orig_size_kb:.1f}KB -> {new_size_kb:.1f}KB)")
            converted += 1
        except Exception as e:
            log(f"    [失败] {filename}: {e}")
            if os.path.exists(temp_path):
                os.remove(temp_path)
    return converted


def discover_sheets():
    """扫描 _data/data.xlsx，返回可用模块列表"""
    xlsx_path = os.path.join(PROJECT_DIR, "_data", "data.xlsx")
    if not os.path.exists(xlsx_path):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        sheets = wb.sheetnames
        wb.close()
        modules = []
        for sn in sheets:
            if sn in MODULE_NAMES:
                modules.append({"id": sn, "label": MODULE_NAMES[sn]})
        return modules
    except Exception:
        return []


def run_update(selected_modules, do_webp, log_func, skip_webp=True):
    """执行数据更新"""
    import importlib.util

    # 图片转 WebP
    if do_webp and HAS_PILLOW:
        log_func("=" * 50)
        log_func("  图片转 WebP 并压缩（max_width=1920, quality=75, method=6）")
        if skip_webp:
            log_func("  跳过已存在的 WebP 文件")
        log_func("=" * 50)
        for folder_name in ["images", "icons"]:
            folder_path = os.path.join(PROJECT_DIR, folder_name)
            log_func(f"  [{folder_name}]")
            count = convert_images_to_webp(folder_path, log_func=log_func, skip_existing_webp=skip_webp)
            log_func(f"  共转换 {count} 张图片")
        log_func("[OK] 图片转 WebP - 完成\n")

    # 数据更新
    if selected_modules:
        gen_path = os.path.join(BASE_DIR, "generate_all.py")
        if not os.path.exists(gen_path):
            log_func("[FAIL] 找不到 generate_all.py")
            return

        # 动态导入 generate_all
        sys.path.insert(0, BASE_DIR)
        import generate_all
        try:
            import openpyxl
            xlsx_path = os.path.join(PROJECT_DIR, "_data", "data.xlsx")
            wb = openpyxl.load_workbook(xlsx_path)
            sheets = wb.sheetnames

            for mod_id in selected_modules:
                log_func("=" * 50)
                log_func(f"  更新 {MODULE_NAMES.get(mod_id, mod_id)}")
                log_func("=" * 50)

                old_stdout = sys.stdout
                sys.stdout = io.StringIO()
                try:
                    if mod_id == "announcements" and "announcements" in sheets:
                        n = generate_all.generate_announcements(wb)
                        log_func(f"  announcements.js - {n} 条公告")
                    elif mod_id == "songs" and "songs" in sheets:
                        n = generate_all.generate_songs(wb)
                        log_func(f"  songs/data.js - {n} 首歌曲")
                    elif mod_id == "lives" and "lives" in sheets:
                        n, mc = generate_all.generate_live(wb)
                        extra = f" ({mc} MC)" if mc else ""
                        log_func(f"  live/data.js - {n} 场演唱会{extra}")
                    elif mod_id == "timeline" and "timeline" in sheets:
                        n = generate_all.generate_timeline(wb)
                        log_func(f"  timeline/data.js - {n} 条事件")
                    elif mod_id == "gallery_images" and "gallery_images" in sheets:
                        n = generate_all.generate_gallery(wb)
                        log_func(f"  gallery/data.js - {n} 张图片")
                finally:
                    output = sys.stdout.getvalue()
                    sys.stdout = old_stdout
                    if output.strip():
                        log_func(output.strip())

                log_func(f"[OK] {MODULE_NAMES.get(mod_id, mod_id)} - 更新完成")

            wb.close()

            # 注入版本号
            log_func("=" * 50)
            v = generate_all.inject_version(log_func=log_func)
            log_func(f"  版本号: {v}")
            log_func("[OK] 版本号注入 - 完成")

        finally:
            sys.path.remove(BASE_DIR)


class App:
    def __init__(self, root):
        self.root = root
        root.title(" wijipedia 数据更新工具")
        root.geometry("560x600")
        root.resizable(True, True)
        root.configure(bg="#0d0d0d")

        # 标题
        header = tk.Label(
            root, text=" wijipedia 数据更新工具",
            font=("Microsoft YaHei", 14, "bold"),
            fg="#ff4d4d", bg="#0d0d0d",
        )
        header.pack(pady=(16, 4))

        sub = tk.Label(
            root, text="选择要更新的模块，点击下方按钮执行",
            font=("Microsoft YaHei", 9),
            fg="#666", bg="#0d0d0d",
        )
        sub.pack(pady=(0, 12))

        # 模块选择区域
        self.modules = discover_sheets()
        self.vars = {}

        frame = tk.Frame(root, bg="#0d0d0d")
        frame.pack(fill="x", padx=24)

        if not self.modules:
            tk.Label(frame, text="未找到 _data/data.xlsx", fg="#999", bg="#0d0d0d",
                     font=("Microsoft YaHei", 10)).pack()
        else:
            for m in self.modules:
                var = tk.BooleanVar(value=True)
                self.vars[m["id"]] = var
                cb = tk.Checkbutton(
                    frame,
                    text=m["label"],
                    variable=var,
                    font=("Microsoft YaHei", 10),
                    fg="#ddd", bg="#0d0d0d",
                    selectcolor="#0d0d0d",
                    activebackground="#0d0d0d",
                    activeforeground="#ff4d4d",
                )
                cb.pack(anchor="w", pady=2)

            # 全选 / 取消
            btn_frame = tk.Frame(root, bg="#0d0d0d")
            btn_frame.pack(pady=(8, 0))

            btn_all = tk.Button(
                btn_frame, text="全选", command=self.select_all,
                font=("Microsoft YaHei", 9), fg="#ddd", bg="#1a1a1a",
                relief="flat", padx=12, pady=4, cursor="hand2",
                activebackground="#2a2a2a", activeforeground="#fff",
            )
            btn_all.pack(side="left", padx=4)

            btn_none = tk.Button(
                btn_frame, text="取消全选", command=self.deselect_all,
                font=("Microsoft YaHei", 9), fg="#ddd", bg="#1a1a1a",
                relief="flat", padx=12, pady=4, cursor="hand2",
                activebackground="#2a2a2a", activeforeground="#fff",
            )
            btn_none.pack(side="left", padx=4)

        # 图片转 WebP 选项
        sep = tk.Frame(root, bg="#2a2a2a", height=1)
        sep.pack(fill="x", padx=24, pady=(12, 8))

        webp_var = tk.BooleanVar(value=False)
        self.webp_var = webp_var

        skip_webp_var = tk.BooleanVar(value=True)
        self.skip_webp_var = skip_webp_var

        if HAS_PILLOW:
            cb_webp = tk.Checkbutton(
                root,
                text="图片转 WebP 并压缩（images + icons 目录）",
                variable=webp_var,
                font=("Microsoft YaHei", 10),
                fg="#ddd", bg="#0d0d0d",
                selectcolor="#0d0d0d",
                activebackground="#0d0d0d",
                activeforeground="#ff4d4d",
            )
            cb_webp.pack(anchor="w", padx=24, pady=2)

            cb_skip = tk.Checkbutton(
                root,
                text="  跳过已存在的 WebP 文件",
                variable=skip_webp_var,
                font=("Microsoft YaHei", 9),
                fg="#999", bg="#0d0d0d",
                selectcolor="#0d0d0d",
                activebackground="#0d0d0d",
                activeforeground="#ff4d4d",
            )
            cb_skip.pack(anchor="w", padx=24, pady=1)
        else:
            lbl_no_pil = tk.Label(
                root,
                text="图片转 WebP（需要安装 Pillow: pip install Pillow）",
                font=("Microsoft YaHei", 9),
                fg="#555", bg="#0d0d0d",
            )
            lbl_no_pil.pack(anchor="w", padx=24, pady=2)

        # 执行按钮
        btn_frame_actions = tk.Frame(root, bg="#0d0d0d")
        btn_frame_actions.pack(pady=(12, 4))

        btn_build = tk.Button(
            btn_frame_actions, text="开始更新", command=self.build_selected,
            font=("Microsoft YaHei", 11, "bold"),
            fg="#fff", bg="#ff4d4d",
            relief="flat", padx=24, pady=8, cursor="hand2",
            activebackground="#ff8080", activeforeground="#fff",
        )
        btn_build.pack(side="left", padx=4)

        btn_scan = tk.Button(
            btn_frame_actions, text="一键更新所有画廊图片", command=self.scan_gallery,
            font=("Microsoft YaHei", 10),
            fg="#ddd", bg="#1a1a1a",
            relief="flat", padx=16, pady=8, cursor="hand2",
            activebackground="#2a2a2a", activeforeground="#fff",
        )
        btn_scan.pack(side="left", padx=4)

        # 输出日志
        self.output = tk.Text(
            root, height=12, font=("Consolas", 9),
            bg="#141414", fg="#aaa",
            relief="flat", padx=10, pady=8,
            highlightthickness=1, highlightcolor="#2a2a2a", highlightbackground="#2a2a2a",
        )
        self.output.pack(fill="both", expand=True, padx=24, pady=(0, 16))
        self.output.insert("end", "就绪。\n")

    def select_all(self):
        for v in self.vars.values():
            v.set(True)

    def deselect_all(self):
        for v in self.vars.values():
            v.set(False)

    def log(self, msg):
        self.output.insert("end", msg + "\n")
        self.output.see("end")
        self.root.update()

    def build_selected(self):
        selected = [m["id"] for m in self.modules if self.vars[m["id"]].get()]
        do_webp = self.webp_var.get() if HAS_PILLOW else False
        skip_webp = self.skip_webp_var.get() if HAS_PILLOW else True

        if not selected and not do_webp:
            messagebox.showwarning("未选择", "请至少选择一个模块或勾选图片转 WebP。")
            return

        self.output.delete("1.0", "end")
        self.log("开始更新...\n")

        try:
            run_update(selected, do_webp, self.log, skip_webp=skip_webp)
        except Exception as e:
            self.log(f"[FAIL] 执行出错: {e}")

        self.log("=" * 50)
        self.log("  更新完毕！请刷新浏览器查看变化。")
        self.log("=" * 50)

    def scan_gallery(self):
        self.output.delete("1.0", "end")
        self.log("开始扫描 images/ 文件夹...\n")
        try:
            sys.path.insert(0, BASE_DIR)
            import generate_all
            try:
                n = generate_all.scan_gallery_images(log_func=self.log)
                self.log("=" * 50)
                if n > 0:
                    self.log(f"  新增 {n} 张图片，已写入 gallery_images sheet。")
                    self.log("  请点击「开始更新」来生成新的 gallery/data.js。")
                else:
                    self.log("  无新增图片。")
                self.log("=" * 50)
            finally:
                sys.path.remove(BASE_DIR)
        except Exception as e:
            self.log(f"[FAIL] 执行出错: {e}")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()