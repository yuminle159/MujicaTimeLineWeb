#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
唯鸡百科 · 统一数据生成工具
=============================
从 _data/data.xlsx 统一生成所有模块的 JS 数据文件。
"""

import os
import re
import json
import sys
from collections import OrderedDict
from datetime import datetime

# 修复 Windows 控制台编码（windowed exe 无控制台则跳过）
if sys.platform == "win32" and sys.stdout is not None:
    import io
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

# =========================== 路径配置 ===========================
if getattr(sys, 'frozen', False):
    # PyInstaller --onefile: 数据文件从 exe 所在目录读取，输出也到 exe 所在目录
    BUNDLE_DIR = sys._MEIPASS
    ROOT = os.path.dirname(sys.executable)
else:
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    ROOT = BUNDLE_DIR

DATA_DIR = os.path.join(ROOT, "_data")
DOCS_DIR = os.path.join(ROOT, "_data", "_docs")
XLSX_PATH = os.path.join(DATA_DIR, "data.xlsx")

# 输出文件路径
OUTPUTS = {
    "announcements": os.path.join(ROOT, "announcements.js"),
    "songs":         os.path.join(ROOT, "songs", "data.js"),
    "live":          os.path.join(ROOT, "live", "data.js"),
    "timeline":      os.path.join(ROOT, "timeline", "data.js"),
    "gallery":       os.path.join(ROOT, "gallery", "data.js"),
}

# 旧 xlsx 文件路径（用于 --init 合并）
OLD_XLSX = {
    "announcements":  os.path.join(ROOT, "announcements.xlsx"),
    "songs":          os.path.join(ROOT, "songs", "data.xlsx"),
    "live":           os.path.join(ROOT, "live", "data.xlsx"),
    "timeline":       os.path.join(ROOT, "timeline", "data.xlsx"),
    "gallery":        os.path.join(ROOT, "gallery", "data.xlsx"),
}


# =========================== 工具函数 ===========================
def js_str(s):
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "")

def normalize_date(raw):
    """统一日期格式为 YYYY/M/D"""
    raw = raw.strip()
    if not raw:
        return raw
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})\s", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})$", raw)
    if m:
        return f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
    return raw

def fix_path(p, module_dir):
    """将 images/ 路径转为相对 HTML 的路径"""
    p = p.strip()
    if not p:
        return p
    if module_dir == "root":
        return p
    if p.startswith("images/") or p.startswith("icons/"):
        return "../" + p
    return p

def read_sheet(wb, sheet_name):
    """读取 sheet 为 dict 列表"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(header):
            if h:
                val = row[i]
                d[h.lower()] = str(val).strip() if val is not None and str(val).strip() != "None" else ""
        if any(v for v in d.values()):
            result.append(d)
    return result


# =========================== 1. 公告 ===========================
def generate_announcements(wb):
    raw = read_sheet(wb, "announcements")
    data = []
    for r in raw:
        pinned = str(r.get("pinned", "")).strip().lower() in ("1", "true", "yes")
        data.append({
            "date": r.get("date", "").strip(),
            "msg": r.get("msg", "").strip(),
            "pinned": pinned
        })
    js = "// 自动生成，请勿手动编辑。运行 generate_all.py 更新\n"
    js += "window.ANNOUNCEMENTS = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    with open(OUTPUTS["announcements"], "w", encoding="utf-8") as f:
        f.write(js)
    return len(data)


# =========================== 2. 歌曲 ===========================
def generate_songs(wb):
    songs_raw = read_sheet(wb, "songs")
    comments_raw = read_sheet(wb, "song_comments")
    live_raw = read_sheet(wb, "song_live_history")

    # 合并 comments
    comments_map = {}
    for cr in comments_raw:
        sn = cr.get("song_name", "")
        if not sn:
            continue
        comments_map.setdefault(sn, []).append({
            "text": cr.get("comment_text", ""),
            "source": cr.get("comment_source", ""),
            "from": cr.get("comment_from", "")
        })

    # 合并 live_history
    live_map = {}
    for lr in live_raw:
        sn = lr.get("song_name", "")
        if not sn:
            continue
        live_map.setdefault(sn, []).append({
            "date": normalize_date(lr.get("live_date", "")),
            "venue": lr.get("live_venue", ""),
            "name": lr.get("live_name", ""),
            "has_video": lr.get("has_video", "").lower() == "yes",
            "video_url": lr.get("video_url", "")
        })

    songs = []
    for s in songs_raw:
        sn = s.get("song_name", "")
        if not sn:
            continue
        songs.append({
            "name": sn,
            "name_jp": s.get("song_name_jp", ""),
            "album": s.get("album", ""),
            "album_year": s.get("album_year", ""),
            "release_date": normalize_date(s.get("release_date", "")),
            "cover": fix_path(s.get("cover", ""), "songs"),
            "type": s.get("type", ""),
            "lyricist": s.get("lyricist", ""),
            "composer": s.get("composer", ""),
            "arranger": s.get("arranger", ""),
            "first_stage": s.get("first_stage", ""),
            "mv_url": s.get("mv_url", ""),
            "lyrics_jp": s.get("lyrics_jp", ""),
            "lyrics_cn": s.get("lyrics_cn", ""),
            "appearances": [a.strip() for a in s.get("appearances", "").split(",") if a.strip()],
            "comments": comments_map.get(sn, []),
            "live_history": live_map.get(sn, [])
        })

    lines = []
    lines.append("// 曲目数据")
    lines.append("// 由 generate_all.py 自动生成，请勿手动修改")
    lines.append("")
    lines.append("const songsData = [")
    for i, song in enumerate(songs):
        lines.append("  {")
        lines.append(f'    name: "{js_str(song["name"])}",')
        lines.append(f'    name_jp: "{js_str(song["name_jp"])}",')
        lines.append(f'    album: "{js_str(song["album"])}",')
        lines.append(f'    album_year: "{js_str(song["album_year"])}",')
        lines.append(f'    release_date: "{js_str(song["release_date"])}",')
        lines.append(f'    cover: "{js_str(song["cover"])}",')
        lines.append(f'    type: "{js_str(song["type"])}",')
        lines.append(f'    lyricist: "{js_str(song["lyricist"])}",')
        lines.append(f'    composer: "{js_str(song["composer"])}",')
        lines.append(f'    arranger: "{js_str(song["arranger"])}",')
        lines.append(f'    first_stage: "{js_str(song["first_stage"])}",')
        lines.append(f'    mv_url: "{js_str(song["mv_url"])}",')
        lines.append(f'    lyrics_jp: "{js_str(song["lyrics_jp"])}",')
        lines.append(f'    lyrics_cn: "{js_str(song["lyrics_cn"])}",')
        lines.append(f'    appearances: {json.dumps(song["appearances"], ensure_ascii=False)},')
        if song["comments"]:
            lines.append("    comments: [")
            for j, c in enumerate(song["comments"]):
                comma = "," if j < len(song["comments"]) - 1 else ""
                lines.append(f'      {{ text: "{js_str(c["text"])}", source: "{js_str(c["source"])}", from: "{js_str(c["from"])}" }}{comma}')
            lines.append("    ],")
        else:
            lines.append("    comments: [],")
        if song["live_history"]:
            lines.append("    live_history: [")
            for j, lh in enumerate(song["live_history"]):
                comma = "," if j < len(song["live_history"]) - 1 else ""
                lines.append(f'      {{ date: "{js_str(lh["date"])}", venue: "{js_str(lh["venue"])}", name: "{js_str(lh["name"])}", has_video: {"true" if lh["has_video"] else "false"}, video_url: "{js_str(lh["video_url"])}" }}{comma}')
            lines.append("    ]")
        else:
            lines.append("    live_history: []")
        lines.append("  }" + ("," if i < len(songs) - 1 else ""))
    lines.append("];")

    with open(OUTPUTS["songs"], "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return len(songs)


# =========================== 3. 演唱会 ===========================
def generate_live(wb):
    lives_raw = read_sheet(wb, "lives")
    setlist_raw = read_sheet(wb, "setlist")
    backstage_raw = read_sheet(wb, "backstage")

    # 合并 setlist
    setlist_map = {}
    for sr in setlist_raw:
        sn = sr.get("live_name", "")
        if not sn:
            continue
        setlist_map.setdefault(sn, []).append({
            "num": sr.get("track_num", ""),
            "title": sr.get("track_title", ""),
            "highlight_label": sr.get("highlight_label", ""),
            "highlight_text": sr.get("highlight_text", ""),
            "mc_file": sr.get("mc_file", ""),
        })

    # 合并 backstage
    backstage_map = {}
    for br in backstage_raw:
        sn = br.get("live_name", "")
        if not sn:
            continue
        backstage_map.setdefault(sn, []).append({
            "photo": fix_path(br.get("photo", ""), "live"),
            "credit": br.get("credit", ""),
            "credit_text": br.get("credit_text", ""),
            "source_url": br.get("source_url", ""),
            "source_label": br.get("source_label", ""),
            "group_id": br.get("group_id", "")
        })

    lives = []
    for l in lives_raw:
        name = l.get("live_name", "")
        if not name:
            continue
        lives.append({
            "name": name,
            "date": normalize_date(l.get("live_date", "")),
            "venue": l.get("live_venue", ""),
            "poster": fix_path(l.get("poster", ""), "live"),
            "kv": fix_path(l.get("kv", ""), "live"),
            "video_url": l.get("video_url", ""),
            "description": l.get("description", ""),
            "setlist": setlist_map.get(name, []),
            "backstage": backstage_map.get(name, [])
        })

    # 加载 MC 内容
    MC_DIR = os.path.join(ROOT, "_data")
    mc_count = 0
    for live in lives:
        for track in live.get("setlist", []):
            mc_file = track.get("mc_file", "").strip()
            if mc_file:
                md_path = os.path.join(MC_DIR, mc_file)
                if os.path.isfile(md_path):
                    with open(md_path, "r", encoding="utf-8") as f:
                        track["mc_content"] = f.read().strip()
                    mc_count += 1

    lines = []
    lines.append("// 演唱会数据")
    lines.append("// 由 generate_all.py 自动生成，请勿手动修改")
    lines.append("")
    lines.append("const livesData = [")
    for i, live in enumerate(lives):
        lines.append("  {")
        lines.append(f'    name: "{js_str(live["name"])}",')
        lines.append(f'    date: "{js_str(live["date"])}",')
        lines.append(f'    venue: "{js_str(live["venue"])}",')
        lines.append(f'    poster: "{js_str(live["poster"])}",')
        lines.append(f'    kv: "{js_str(live["kv"])}",')
        lines.append(f'    video_url: "{js_str(live["video_url"])}",')
        lines.append(f'    description: "{js_str(live["description"])}",')
        lines.append(f'    setlist: {json.dumps(live["setlist"], ensure_ascii=False)},')
        lines.append(f'    backstage: {json.dumps(live["backstage"], ensure_ascii=False)}')
        lines.append("  }" + ("," if i < len(lives) - 1 else ""))
    lines.append("];")

    with open(OUTPUTS["live"], "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return len(lives), mc_count


# =========================== 4. 时间轴 ===========================
def generate_timeline(wb):
    raw = read_sheet(wb, "timeline")
    if not raw:
        return 0

    # 构建列映射
    header_keys = list(raw[0].keys())
    data_rows = []
    for d in raw:
        data_rows.append([d.get(k, "") for k in header_keys])

    # 构建列索引
    col_map = {name.strip().lower(): i for i, name in enumerate(header_keys)}
    def get_col(row, name):
        idx = col_map.get(name.strip().lower())
        if idx is not None and idx < len(row):
            return (row[idx] or "").strip()
        return ""

    events = OrderedDict()
    grouped_rows = OrderedDict()

    for row in data_rows:
        group = get_col(row, "group")
        date = normalize_date(get_col(row, "date"))
        title = get_col(row, "title")
        category = get_col(row, "category")
        desc = get_col(row, "description")
        tag = get_col(row, "tag")
        mt = get_col(row, "media_type").lower()

        if group:
            grouped_rows.setdefault(group, []).append({
                "date": date, "title": title, "category": category,
                "desc": desc, "tag": tag, "media_type": mt,
                "media_src": get_col(row, "media_src"),
                "media_caption": get_col(row, "media_caption"),
                "media_url": get_col(row, "media_url"),
                "media_title": get_col(row, "media_title"),
            })
            continue

        key = (date, title, category, desc, tag)
        events.setdefault(key, [])
        media = None
        if mt == "image":
            media = {"type": "image", "src": fix_path(get_col(row, "media_src"), "timeline")}
            if get_col(row, "media_caption"):
                media["caption"] = get_col(row, "media_caption")
        elif mt == "video":
            media = {"type": "video", "src": get_col(row, "media_src")}
            if get_col(row, "media_caption"):
                media["caption"] = get_col(row, "media_caption")
        elif mt == "link":
            media = {"type": "link", "url": get_col(row, "media_url")}
            if get_col(row, "media_title"):
                media["title"] = get_col(row, "media_title")
        if media:
            events[key].append(media)

    # 处理分组事件
    for group_name, rows in grouped_rows.items():
        dates = sorted(set(r["date"] for r in rows if r["date"]))
        merged_date = dates[0] if len(dates) == 1 else f"{dates[0]} - {dates[-1]}"
        first = rows[0]
        key = (merged_date, first["title"], first["category"], first["desc"], first["tag"])
        events[key] = []
        for r in rows:
            mt = r["media_type"].lower()
            media = None
            if mt == "image":
                media = {"type": "image", "src": fix_path(r["media_src"], "timeline")}
                if r["media_caption"]:
                    media["caption"] = r["media_caption"]
            elif mt == "video":
                media = {"type": "video", "src": r["media_src"]}
                if r["media_caption"]:
                    media["caption"] = r["media_caption"]
            elif mt == "link":
                media = {"type": "link", "url": r["media_url"]}
                if r["media_title"]:
                    media["title"] = r["media_title"]
            if media:
                events[key].append(media)

    lines = []
    lines.append("// 时间轴数据")
    lines.append("// 由 generate_all.py 自动生成，请勿手动修改")
    lines.append("")
    lines.append("const timelineData = [")
    event_items = list(events.items())
    for i, (key, media_list) in enumerate(event_items):
        date, title, category, desc, tag = key
        lines.append("  {")
        lines.append(f'    date: "{js_str(date)}",')
        lines.append(f'    title: "{js_str(title)}",')
        lines.append(f'    category: "{js_str(category)}",')
        lines.append(f'    description: "{js_str(desc)}",')
        lines.append(f'    tag: "{js_str(tag)}",')
        if media_list:
            lines.append("    media: [")
            for j, m in enumerate(media_list):
                if m["type"] == "image":
                    cap = f', caption: "{js_str(m["caption"])}"' if "caption" in m else ""
                    lines.append(f'      {{ type: "image", src: "{js_str(m["src"])}"{cap} }}' + ("," if j < len(media_list) - 1 else ""))
                elif m["type"] == "video":
                    cap = f', caption: "{js_str(m["caption"])}"' if "caption" in m else ""
                    lines.append(f'      {{ type: "video", src: "{js_str(m["src"])}"{cap} }}' + ("," if j < len(media_list) - 1 else ""))
                elif m["type"] == "link":
                    ttl = f', title: "{js_str(m["title"])}"' if "title" in m else ""
                    lines.append(f'      {{ type: "link", url: "{js_str(m["url"])}"{ttl} }}' + ("," if j < len(media_list) - 1 else ""))
            lines.append("    ]")
        else:
            lines.append("    media: []")
        lines.append("  }" + ("," if i < len(event_items) - 1 else ""))
    lines.append("];")
    lines.append("")
    lines.append("const timelineConfig = {")
    lines.append('  zeroDate: "2023-06-04",')
    lines.append("  pixelsPerDay: 4")
    lines.append("};")

    with open(OUTPUTS["timeline"], "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return len(event_items)


# =========================== 5. 画廊 ===========================
def parse_gallery_tags(raw_tags):
    """解析 "LIVE:0th, LIVE:1st" 为嵌套 dict"""
    result = {}
    if not raw_tags:
        return result
    for tag in raw_tags.split(","):
        tag = tag.strip()
        if not tag:
            continue
        if ":" in tag:
            cat, sub = tag.split(":", 1)
            cat, sub = cat.strip(), sub.strip()
            if sub:
                result.setdefault(cat, [])
                if sub not in result[cat]:
                    result[cat].append(sub)
        else:
            cat = tag.strip()
            if cat and cat not in result:
                result[cat] = []
    return result

def generate_gallery(wb):
    images_raw = read_sheet(wb, "gallery_images")
    images = []
    for img in images_raw:
        if not img.get("filename"):
            continue
        images.append({
            "filename": img.get("filename", ""),
            "title": img.get("title", ""),
            "date": img.get("date", ""),
            "tags": parse_gallery_tags(img.get("tags", "")),
            "description": img.get("description", "")
        })

    lines = []
    lines.append("// 画廊图片数据")
    lines.append("// 由 generate_all.py 自动生成，请勿手动修改")
    lines.append("")
    lines.append("const galleryData = [")
    for i, item in enumerate(images):
        lines.append("  {")
        lines.append(f'    filename: "{js_str(item["filename"])}",')
        lines.append(f'    title: "{js_str(item["title"])}",')
        lines.append(f'    date: "{js_str(item["date"])}",')
        lines.append(f'    tags: {json.dumps(item["tags"], ensure_ascii=False)},')
        lines.append(f'    description: "{js_str(item["description"])}"')
        lines.append("  }" + ("," if i < len(images) - 1 else ""))
    lines.append("];")

    with open(OUTPUTS["gallery"], "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return len(images)


# =========================== 初始化：合并旧 xlsx ===========================
def init_merged_xlsx():
    """从旧的分散 xlsx 合并创建 _data/data.xlsx"""
    os.makedirs(DATA_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    # 定义 sheet 名称映射：旧文件 key → (sheet 名称, 源 sheet 名称)
    sheet_sources = {
        "announcements":     [("announcements", None)],
        "songs":             [("songs", "songs"), ("song_comments", "comments"), ("song_live_history", "live_history")],
        "live":              [("lives", "lives"), ("setlist", "setlist"), ("backstage", "backstage")],
        "gallery":           [("gallery_images", "images")],
        "timeline":          [("timeline", "时间轴数据")],
    }

    merged_count = 0
    for module_key, sheet_list in sheet_sources.items():
        old_path = OLD_XLSX[module_key]
        if not os.path.exists(old_path):
            print(f"  [SKIP] 旧文件不存在: {old_path}")
            continue
        old_wb = openpyxl.load_workbook(old_path)
        for new_sheet_name, old_sheet_name in sheet_list:
            src_name = old_sheet_name or old_wb.sheetnames[0] if old_wb.sheetnames else None
            if src_name and src_name in old_wb.sheetnames:
                ws_src = old_wb[src_name]
                ws_new = wb.create_sheet(new_sheet_name)
                for row in ws_src.iter_rows(values_only=True):
                    ws_new.append(list(row))
                merged_count += 1
                print(f"  [OK] {module_key}/{src_name} -> {new_sheet_name}")
        old_wb.close()

    wb.save(XLSX_PATH)
    print(f"\n合并完成！{merged_count} 个 Sheet 已写入 {XLSX_PATH}")
    return merged_count


# =========================== 主流程 ===========================
def inject_version(log_func=None):
    """在所有 HTML 中为 data.js / announcements.js 引用注入版本号，输出版本号字符串"""
    version = datetime.now().strftime("%Y%m%d%H%M")
    html_files = [
        os.path.join(ROOT, "index.html"),
        os.path.join(ROOT, "songs", "index.html"),
        os.path.join(ROOT, "live", "index.html"),
        os.path.join(ROOT, "timeline", "index.html"),
        os.path.join(ROOT, "gallery", "index.html"),
    ]
    # 匹配所有本地 .css / .js / .svg 引用（跳过 https:// 外部链接）
    pattern = re.compile(
        r'((?:href|src)="(?!https?://)(?!data:)[^"]*\.(?:css|js|svg))'
        r'(?:\?v=[^"]*)?(")'
    )

    for html_path in html_files:
        if not os.path.exists(html_path):
            continue
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
        new_html = pattern.sub(rf"\1?v={version}\2", html)
        if new_html != html:
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(new_html)
            (log_func or print)(f"  ✓ {os.path.relpath(html_path, ROOT)} → v={version}")
    return version


def main():
    if len(sys.argv) > 1 and sys.argv[1] == "--init":
        print("=== 初始化：从旧 xlsx 合并创建 _data/data.xlsx ===")
        init_merged_xlsx()
        print("\n现在可以运行 py generate_all.py 来生成所有 JS 文件。")
        return

    if not os.path.exists(XLSX_PATH):
        print(f"错误: 找不到 {XLSX_PATH}")
        print("请先运行 py generate_all.py --init 来从旧文件合并创建。")
        sys.exit(1)

    print("=== 唯鸡百科 · 统一数据生成 ===")
    print(f"数据源: {XLSX_PATH}\n")

    wb = openpyxl.load_workbook(XLSX_PATH)
    sheets = wb.sheetnames
    print(f"Sheet 列表: {', '.join(sheets)}\n")

    results = {}

    # 公告
    if "announcements" in sheets:
        n = generate_announcements(wb)
        results["公告"] = f"{n} 条"
        print(f"  ✓ announcements.js — {n} 条公告")

    # 歌曲
    if "songs" in sheets:
        n = generate_songs(wb)
        results["歌曲"] = f"{n} 首"
        print(f"  ✓ songs/data.js — {n} 首歌曲")

    # 演唱会
    if "lives" in sheets:
        n, mc = generate_live(wb)
        extra = f" ({mc} MC)" if mc else ""
        results["演唱会"] = f"{n} 场{extra}"
        print(f"  ✓ live/data.js — {n} 场演唱会" + (f"，{mc} 个 MC 文件" if mc else ""))

    # 时间轴
    if "timeline" in sheets:
        n = generate_timeline(wb)
        results["时间轴"] = f"{n} 条事件"
        print(f"  ✓ timeline/data.js — {n} 条事件")

    # 画廊
    if "gallery_images" in sheets:
        n = generate_gallery(wb)
        results["画廊"] = f"{n} 张图片"
        print(f"  ✓ gallery/data.js — {n} 张图片")

    wb.close()

    print("\n--- 注入版本号 ---")
    v = inject_version()
    print(f"  版本号: {v}")

    print("\n=== 全部完成 ===")
    for k, v in results.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()